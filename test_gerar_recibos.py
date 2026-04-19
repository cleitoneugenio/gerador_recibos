"""
Suite de testes para gerar_recibos.py.
Cobre: fmt_valor, valor_por_extenso, formatar_data, ler_funcionarios,
       gerar_pdf_de_lista, gerar_preview_pil, resource_path.
"""
import os
import sys
import tempfile
import threading
from datetime import datetime

import openpyxl
import pytest

from gerar_recibos import (
    fmt_valor,
    formatar_data,
    gerar_pdf_de_lista,
    gerar_preview_pil,
    ler_funcionarios,
    resource_path,
    valor_por_extenso,
)


# ── Fixtures ─────────────────────────────────────────────────────────────────

@pytest.fixture
def xlsx_simples(tmp_path):
    """Planilha com 2 funcionários válidos + linha TOTAL."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(['Num', 'Nome', 'c2', 'c3', 'c4', 'c5', 'c6', 'c7', 'c8', 'c9', 'Total'])
    ws.append([1, 'Maria Souza',  '', '', '', '', '', '', '', '', 1500.0])
    ws.append([2, 'Joao Silva',   '', '', '', '', '', '', '', '', 2000.5])
    ws.append(['TOTAL', '',       '', '', '', '', '', '', '', '', 3500.5])
    path = tmp_path / 'planilha.xlsx'
    wb.save(path)
    return str(path)


@pytest.fixture
def xlsx_vazia(tmp_path):
    """Planilha sem funcionários — só cabeçalho."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(['Num', 'Nome', 'c2', 'c3', 'c4', 'c5', 'c6', 'c7', 'c8', 'c9', 'Total'])
    path = tmp_path / 'vazia.xlsx'
    wb.save(path)
    return str(path)


@pytest.fixture
def xlsx_invalidos(tmp_path):
    """Planilha com totais zero/negativos — todos devem ser ignorados."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(['Num', 'Nome', 'c2', 'c3', 'c4', 'c5', 'c6', 'c7', 'c8', 'c9', 'Total'])
    ws.append([1, 'Sem Servico', '', '', '', '', '', '', '', '', 0.0])
    ws.append([2, 'Negativo',    '', '', '', '', '', '', '', '', -100.0])
    path = tmp_path / 'invalidos.xlsx'
    wb.save(path)
    return str(path)


# ── fmt_valor ─────────────────────────────────────────────────────────────────

class TestFmtValor:
    def test_valor_tipico(self):
        assert fmt_valor(1234.56) == 'R$ 1.234,56'

    def test_zero(self):
        assert fmt_valor(0.0) == 'R$ 0,00'

    def test_milhao(self):
        assert fmt_valor(1_000_000.0) == 'R$ 1.000.000,00'

    def test_centavo_minimo(self):
        assert fmt_valor(0.01) == 'R$ 0,01'

    def test_sem_centavos(self):
        assert fmt_valor(500.0) == 'R$ 500,00'


# ── valor_por_extenso ─────────────────────────────────────────────────────────

class TestValorPorExtenso:
    def test_reais_e_centavos(self):
        result = valor_por_extenso(1500.50)
        assert 'mil' in result
        assert 'centavos' in result

    def test_sem_centavos(self):
        result = valor_por_extenso(100.0)
        assert 'centavos' not in result

    def test_contem_formatacao_monetaria(self):
        result = valor_por_extenso(250.0)
        assert 'R$' in result
        assert '(' in result and ')' in result

    def test_negativo_nao_explode(self):
        # Valores negativos não são bloqueados na lógica de negócio — apenas
        # ler_funcionarios os filtra. Verificamos que não lança exceção.
        result = valor_por_extenso(-50.0)
        assert isinstance(result, str)


# ── formatar_data ─────────────────────────────────────────────────────────────

class TestFormatarData:
    def test_formato_pt_br(self):
        assert formatar_data(datetime(2025, 1, 5)) == '5 de janeiro de 2025'

    def test_todos_os_meses(self):
        meses = [
            'janeiro', 'fevereiro', 'março', 'abril', 'maio', 'junho',
            'julho', 'agosto', 'setembro', 'outubro', 'novembro', 'dezembro',
        ]
        for m_idx, nome in enumerate(meses, start=1):
            result = formatar_data(datetime(2025, m_idx, 1))
            assert nome in result, f'Mês {m_idx} incorreto: {result}'

    def test_dia_sem_zero(self):
        result = formatar_data(datetime(2025, 3, 7))
        assert result.startswith('7 ')


# ── ler_funcionarios ──────────────────────────────────────────────────────────

class TestLerFuncionarios:
    def test_carrega_funcionarios_validos(self, xlsx_simples):
        result = ler_funcionarios(xlsx_simples)
        assert len(result) == 2
        assert result[0]['nome'] == 'Maria Souza'
        assert result[0]['total'] == 1500.0
        assert result[1]['nome'] == 'Joao Silva'
        assert result[1]['total'] == 2000.5

    def test_planilha_vazia(self, xlsx_vazia):
        assert ler_funcionarios(xlsx_vazia) == []

    def test_total_zero_ignorado(self, xlsx_invalidos):
        assert ler_funcionarios(xlsx_invalidos) == []

    def test_linha_total_ignorada(self, xlsx_simples):
        result = ler_funcionarios(xlsx_simples)
        nomes = [f['nome'] for f in result]
        assert 'TOTAL' not in nomes
        assert '' not in nomes

    def test_arquivo_inexistente(self):
        with pytest.raises(FileNotFoundError):
            ler_funcionarios('nao_existe.xlsx')


# ── gerar_pdf_de_lista ────────────────────────────────────────────────────────

class TestGerarPdfDeLista:
    def test_gera_pdf_valido(self, tmp_path):
        funcs = [{'nome': 'Maria Souza', 'total': 1500.0}]
        pdf = str(tmp_path / 'out.pdf')
        gerar_pdf_de_lista(funcs, pdf)
        assert os.path.exists(pdf)
        assert os.path.getsize(pdf) > 1000

    def test_lista_vazia_lanca_value_error(self, tmp_path):
        with pytest.raises(ValueError, match='Nenhum funcionário'):
            gerar_pdf_de_lista([], str(tmp_path / 'out.pdf'))

    def test_multiplos_funcionarios(self, tmp_path):
        funcs = [{'nome': f'Func {i}', 'total': float(i * 100)} for i in range(1, 11)]
        pdf = str(tmp_path / 'out.pdf')
        gerar_pdf_de_lista(funcs, pdf)
        assert os.path.getsize(pdf) > 5000

    def test_nome_com_caracteres_xml(self, tmp_path):
        """xml_escape deve proteger contra injection no PDF."""
        funcs = [
            {'nome': '<script>alert(1)</script>', 'total': 100.0},
            {'nome': 'Maria & João "Teste"', 'total': 200.0},
        ]
        pdf = str(tmp_path / 'out.pdf')
        gerar_pdf_de_lista(funcs, pdf)
        assert os.path.exists(pdf)

    def test_pdf_cresce_com_volume(self, tmp_path):
        def make_pdf(n):
            funcs = [{'nome': f'Func {i}', 'total': 100.0} for i in range(n)]
            pdf = str(tmp_path / f'out_{n}.pdf')
            gerar_pdf_de_lista(funcs, pdf)
            return os.path.getsize(pdf)

        sz_10 = make_pdf(10)
        sz_50 = make_pdf(50)
        assert sz_50 > sz_10


# ── gerar_preview_pil ─────────────────────────────────────────────────────────

class TestGerarPreviewPil:
    def test_retorna_imagem(self):
        from PIL import Image
        func = {'nome': 'Maria Souza', 'total': 1500.0}
        img = gerar_preview_pil(func, width=420)
        assert isinstance(img, Image.Image)
        assert img.width == 420

    def test_width_minimo(self):
        img = gerar_preview_pil({'nome': 'X', 'total': 1.0}, width=100)
        assert img.width == 100

    def test_caracteres_especiais_no_nome(self):
        func = {'nome': 'Ana & João "Teste" <OK>', 'total': 99.99}
        img = gerar_preview_pil(func, width=420)
        assert img.width == 420

    def test_valor_alto(self):
        img = gerar_preview_pil({'nome': 'Rico', 'total': 9_999_999.99}, width=420)
        assert img.width == 420

    def test_thread_safety(self):
        """Múltiplas threads renderizando simultaneamente não devem gerar erros."""
        results, errors = [], []

        def render(i):
            try:
                img = gerar_preview_pil({'nome': f'Func {i}', 'total': float(i * 100)}, width=320)
                results.append(img.size)
            except Exception as exc:
                errors.append(str(exc))

        threads = [threading.Thread(target=render, args=(i,)) for i in range(8)]
        for t in threads:
            t.start()
        for t in threads:
            t.join()

        assert errors == [], f'Erros em threads: {errors}'
        assert len(results) == 8


# ── resource_path ─────────────────────────────────────────────────────────────

class TestResourcePath:
    def test_fora_de_pyinstaller(self):
        p = resource_path('recibo.ico')
        assert p.endswith('recibo.ico')
        assert os.path.dirname(p) == os.path.dirname(os.path.abspath('gerar_recibos.py'))

    def test_dentro_de_pyinstaller(self, monkeypatch):
        monkeypatch.setattr(sys, '_MEIPASS', '/tmp/fake_meipass', raising=False)
        p = resource_path('recibo.ico')
        assert p == os.path.join('/tmp/fake_meipass', 'recibo.ico')
